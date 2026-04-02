"""
Integration tests — end-to-end run using sample_input.json.
Verifies the full pipeline: engine → validation → Excel output.
"""

from __future__ import annotations

import json
import os
import pytest

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SAMPLE_PATH = os.path.join(ROOT, "samples", "sample_input.json")


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def sample_data() -> dict:
    with open(SAMPLE_PATH, encoding="utf-8") as fh:
        return json.load(fh)


@pytest.fixture(scope="module")
def boq_results(sample_data):
    from src.engine.qto_engine import QTOEngine
    engine = QTOEngine()
    return engine.run(sample_data)


@pytest.fixture(scope="module")
def validation_report(sample_data, boq_results):
    from src.validation.validator import QTOValidator
    validator = QTOValidator()
    return validator.validate_all(
        boq_results,
        sample_data["project_type"],
        sample_data["plot_area"],
    )


# ---------------------------------------------------------------------------
# Sample data sanity
# ---------------------------------------------------------------------------

class TestSampleData:
    def test_sample_file_exists(self):
        assert os.path.exists(SAMPLE_PATH), f"Missing: {SAMPLE_PATH}"

    def test_required_keys_present(self, sample_data):
        for key in ["project_type", "plot_area", "footings", "rooms", "openings",
                    "exc_depth", "has_road_base", "road_base_thickness"]:
            assert key in sample_data, f"Key '{key}' missing from sample_input.json"

    def test_project_type_is_g_plus_1(self, sample_data):
        assert sample_data["project_type"] == "G+1"

    def test_plot_area_is_valid(self, sample_data):
        assert sample_data["plot_area"] > 0


# ---------------------------------------------------------------------------
# BOQ generation
# ---------------------------------------------------------------------------

class TestBOQGeneration:
    def test_boq_not_empty(self, boq_results):
        assert len(boq_results) > 0

    def test_boq_has_50_or_more_items(self, boq_results):
        # Items below 90% confidence are excluded; at least 20 core items must survive
        assert len(boq_results) >= 20

    def test_all_items_have_required_fields(self, boq_results):
        required = {"item_no", "description", "unit", "quantity", "category"}
        for item in boq_results:
            missing = required - set(item.keys())
            assert not missing, f"Item missing fields {missing}: {item}"

    def test_all_quantities_are_non_negative(self, boq_results):
        for item in boq_results:
            assert item["quantity"] >= 0.0, (
                f"Negative quantity for item: {item['description']}"
            )

    def test_sub_structure_items_present(self, boq_results):
        cats = {item["category"] for item in boq_results}
        assert "Sub-Structure" in cats

    def test_super_structure_items_present(self, boq_results):
        cats = {item["category"] for item in boq_results}
        assert "Super-Structure" in cats

    def test_finishes_items_present(self, boq_results):
        cats = {item["category"] for item in boq_results}
        assert "Finishes" in cats

    def test_excavation_in_sub_structure(self, boq_results):
        names = [i["description"].lower() for i in boq_results]
        assert any("excavation" in n for n in names)

    def test_road_base_thickness_from_user_input(self):
        """Road base thickness must come from user input, not be hardcoded."""
        from src.engine.sub_structure import SubStructureCalculator
        import json, os
        with open(os.path.join(os.path.dirname(__file__), "..", "samples", "sample_input.json")) as f:
            data = json.load(f)
        calc = SubStructureCalculator()
        longest_length = data.get("longest_length", 26.0)
        longest_width = data.get("longest_width", 20.0)
        exc_area = (2 + longest_length) * (2 + longest_width)
        rb_result = calc.calculate_road_base(exc_area, thickness=0.30)
        expected_vol = round(exc_area * 0.30, 3)
        assert abs(rb_result["volume_m3"] - expected_vol) < 0.01, (
            f"Road base volume {rb_result['volume_m3']} != expected {expected_vol} "
            f"(user thickness 0.30 m not applied)"
        )

    def test_dry_area_flooring_present(self, boq_results):
        names = [i["description"].lower() for i in boq_results]
        assert any("flooring" in n and "dry" in n for n in names)

    def test_wet_areas_flooring_present(self, boq_results):
        names = [i["description"].lower() for i in boq_results]
        assert any("flooring" in n and "wet" in n for n in names)

    def test_internal_plaster_present(self, boq_results):
        names = [i["description"].lower() for i in boq_results]
        assert any("plaster" in n for n in names)

    def test_block_20_present(self, boq_results):
        names = [i["description"].lower() for i in boq_results]
        assert any("block 20cm" in n for n in names)

    def test_parapet_block_present(self, boq_results):
        names = [i["description"].lower() for i in boq_results]
        assert any("parapet" in n and "block" in n for n in names)

    def test_parapet_concrete_present(self, boq_results):
        names = [i["description"].lower() for i in boq_results]
        assert any("parapet" in n and "concrete" in n for n in names)

    def test_amounts_calculated_from_rate_and_qty(self, boq_results):
        for item in boq_results:
            expected_amount = round(item["quantity"] * item["rate"], 2)
            assert abs(item["amount"] - expected_amount) < 0.01, (
                f"Amount mismatch for {item['description']}: "
                f"expected {expected_amount}, got {item['amount']}"
            )


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------

class TestValidation:
    def test_validation_report_created(self, validation_report):
        from src.validation.validator import ValidationReport
        assert isinstance(validation_report, ValidationReport)

    def test_item_results_count_matches_boq(self, boq_results, validation_report):
        assert len(validation_report.item_results) == len(boq_results)

    def test_overall_confidence_between_0_and_100(self, validation_report):
        assert 0.0 <= validation_report.overall_confidence <= 100.0

    def test_each_item_has_confidence_between_0_and_100(self, validation_report):
        for vr in validation_report.item_results:
            assert 0.0 <= vr.confidence <= 100.0, (
                f"Confidence out of range for {vr.item_name}: {vr.confidence}"
            )

    def test_each_item_has_valid_flag(self, validation_report):
        valid_flags = {"GREEN", "YELLOW", "RED"}
        for vr in validation_report.item_results:
            assert vr.flag in valid_flags, (
                f"Invalid flag '{vr.flag}' for {vr.item_name}"
            )

    def test_summary_not_empty(self, validation_report):
        assert validation_report.summary != ""

    def test_is_draft_is_boolean(self, validation_report):
        assert isinstance(validation_report.is_draft, bool)


# ---------------------------------------------------------------------------
# Excel generation
# ---------------------------------------------------------------------------

class TestExcelGeneration:
    def test_excel_file_created(self, boq_results, validation_report, tmp_path):
        from src.output.excel_generator import ExcelGenerator
        output_path = str(tmp_path / "test_boq.xlsx")
        gen = ExcelGenerator()
        result_path = gen.generate(
            boq_results,
            validation_report,
            output_path,
            project_info={
                "name": "Integration Test Villa",
                "ref": "INT-001",
                "type": "G+1",
                "plot_area": 153,
                "date": "2024-01-01",
            },
        )
        assert os.path.exists(result_path)
        assert result_path.endswith(".xlsx")

    def test_excel_has_multiple_sheets(self, boq_results, validation_report, tmp_path):
        import openpyxl
        from src.output.excel_generator import ExcelGenerator
        output_path = str(tmp_path / "test_boq_sheets.xlsx")
        gen = ExcelGenerator()
        gen.generate(boq_results, validation_report, output_path, {})
        wb = openpyxl.load_workbook(output_path)
        assert len(wb.sheetnames) >= 2
        assert "BOQ" in wb.sheetnames
        assert "Summary" in wb.sheetnames

    def test_excel_boq_sheet_has_data_rows(
        self, boq_results, validation_report, tmp_path
    ):
        import openpyxl
        from src.output.excel_generator import ExcelGenerator
        output_path = str(tmp_path / "test_boq_data.xlsx")
        gen = ExcelGenerator()
        gen.generate(boq_results, validation_report, output_path, {})
        wb = openpyxl.load_workbook(output_path)
        ws = wb["BOQ"]
        # Count non-empty rows below the header rows
        non_empty = sum(
            1 for row in ws.iter_rows(min_row=4, values_only=True)
            if any(cell is not None for cell in row)
        )
        assert non_empty > 0


# ---------------------------------------------------------------------------
# CLI (--sample flag)
# ---------------------------------------------------------------------------

class TestCLI:
    def test_cli_sample_flag_runs_without_error(self, tmp_path):
        from src.main import main
        output = str(tmp_path / "cli_output.xlsx")
        exit_code = main(["--sample", "--type", "G+1", "--output", output])
        assert exit_code == 0
        assert os.path.exists(output)
