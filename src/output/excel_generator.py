"""
Excel Generator — produces a professional BOQ workbook with colour-coded
validation results, section totals, and a summary sheet.
"""

from __future__ import annotations

import os
from datetime import datetime
from typing import Any

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
    )
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------

COLOR_HEADER_BG = "1F4E79"    # dark blue
COLOR_HEADER_FG = "FFFFFF"    # white

COLOR_SECTION_BG = "D6E4F0"   # light blue for section headings
COLOR_SECTION_FG = "1F4E79"

COLOR_GREEN_BG = "C6EFCE"
COLOR_GREEN_FG = "276221"

COLOR_YELLOW_BG = "FFEB9C"
COLOR_YELLOW_FG = "9C5700"

COLOR_RED_BG = "FFC7CE"
COLOR_RED_FG = "9C0006"

COLOR_TOTAL_BG = "E2EFDA"
COLOR_TOTAL_FG = "276221"

COLOR_ESTIMATED_BG = "E2D0F7"   # soft purple — statistical estimate
COLOR_ESTIMATED_FG = "5B0099"

COLOR_DRAFT_BG = "FFC7CE"
COLOR_FINAL_BG = "C6EFCE"


def _fill(hex_color: str) -> "PatternFill":
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _font(bold: bool = False, color: str = "000000", size: int = 10) -> "Font":
    return Font(bold=bold, color=color, size=size, name="Calibri")


def _border() -> "Border":
    side = Side(style="thin")
    return Border(left=side, right=side, top=side, bottom=side)


def _center() -> "Alignment":
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left() -> "Alignment":
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------
# Generator
# ---------------------------------------------------------------------------

class ExcelGenerator:
    """Generates a professional BOQ Excel workbook."""

    COLUMNS = [
        ("Item No", 10),
        ("Description", 50),
        ("Unit", 8),
        ("Quantity", 12),
        ("Rate (AED)", 14),
        ("Amount (AED)", 16),
        ("Confidence %", 14),
        ("Flag", 10),
        ("Notes", 40),
    ]

    def generate(
        self,
        boq_results: list[dict],
        validation_report: Any,          # ValidationReport dataclass
        output_path: str,
        project_info: dict | None = None,
    ) -> str:
        """
        Generate the BOQ Excel file.

        Parameters
        ----------
        boq_results      : List of BOQ line items from QTOEngine.run()
        validation_report: ValidationReport from QTOValidator.validate_all()
        output_path      : Destination .xlsx path
        project_info     : Optional dict with keys: name, ref, date, type, plot_area

        Returns
        -------
        str: Absolute path to the generated file.
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required: pip install openpyxl")

        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        project_info = project_info or {}
        val_map = self._build_validation_map(validation_report)

        self._create_boq_sheet(wb, boq_results, val_map, project_info)
        self._create_summary_sheet(wb, boq_results, validation_report, project_info)
        self._create_analysis_sheet(wb, boq_results, project_info)

        # Ensure output directory exists
        os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
        wb.save(output_path)
        return os.path.abspath(output_path)

    # ------------------------------------------------------------------
    # BOQ Sheet
    # ------------------------------------------------------------------
    def _create_boq_sheet(
        self,
        wb: "Workbook",
        boq: list[dict],
        val_map: dict,
        info: dict,
    ) -> None:
        ws = wb.create_sheet("BOQ")

        # --- Project info header ---
        self._write_project_header(ws, info)

        # --- Column headers ---
        header_row = ws.max_row + 2
        for col_idx, (col_name, col_width) in enumerate(self.COLUMNS, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=col_name)
            cell.fill = _fill(COLOR_HEADER_BG)
            cell.font = _font(bold=True, color=COLOR_HEADER_FG, size=11)
            cell.alignment = _center()
            cell.border = _border()
            ws.column_dimensions[get_column_letter(col_idx)].width = col_width

        ws.row_dimensions[header_row].height = 30
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

        # --- Data rows grouped by category ---
        current_category = None
        data_start = header_row + 1
        row = data_start

        category_start_rows: dict[str, int] = {}
        section_amounts: dict[str, float] = {}

        for item in boq:
            cat = item.get("category", "General")

            # Section heading
            if cat != current_category:
                current_category = cat
                category_start_rows[cat] = row
                section_amounts.setdefault(cat, 0.0)

                heading_cell = ws.cell(row=row, column=1, value=cat.upper())
                ws.merge_cells(
                    start_row=row, start_column=1,
                    end_row=row, end_column=len(self.COLUMNS)
                )
                heading_cell.fill = _fill(COLOR_SECTION_BG)
                heading_cell.font = _font(bold=True, color=COLOR_SECTION_FG, size=11)
                heading_cell.alignment = _left()
                heading_cell.border = _border()
                ws.row_dimensions[row].height = 20
                row += 1

            # Validation data
            vr = val_map.get(item.get("description", ""))
            confidence = vr.confidence if vr else 100.0
            flag = vr.flag if vr else "GREEN"
            note = vr.note if vr else item.get("confidence_note", "")

            bg = self._flag_bg(flag)
            fg = self._flag_fg(flag)

            values = [
                item.get("item_no", ""),
                item.get("description", ""),
                item.get("unit", ""),
                item.get("quantity", 0.0),
                item.get("rate", 0.0),
                item.get("amount", 0.0),
                confidence,
                flag,
                note,
            ]

            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.border = _border()
                cell.alignment = _center() if col_idx not in (2, 9) else _left()
                # Apply confidence colour to Confidence % and Flag columns only
                if col_idx in (7, 8):
                    cell.fill = _fill(bg)
                    cell.font = _font(color=fg)
                else:
                    cell.font = _font()

                if col_idx == 4:  # Quantity
                    cell.number_format = "#,##0.000"
                elif col_idx in (5, 6):  # Rate / Amount
                    cell.number_format = "#,##0.00"
                elif col_idx == 7:  # Confidence
                    cell.number_format = "0.0"

            section_amounts[cat] = section_amounts.get(cat, 0.0) + item.get("amount", 0.0)
            row += 1

        # Section totals
        for cat, amount in section_amounts.items():
            total_label = ws.cell(row=row, column=1, value=f"Sub-Total — {cat}")
            ws.merge_cells(
                start_row=row, start_column=1,
                end_row=row, end_column=5
            )
            total_label.fill = _fill(COLOR_TOTAL_BG)
            total_label.font = _font(bold=True, color=COLOR_TOTAL_FG)
            total_label.alignment = _left()
            total_label.border = _border()

            total_amount = ws.cell(row=row, column=6, value=amount)
            total_amount.fill = _fill(COLOR_TOTAL_BG)
            total_amount.font = _font(bold=True, color=COLOR_TOTAL_FG)
            total_amount.number_format = "#,##0.00"
            total_amount.border = _border()
            row += 1

        # Grand total
        grand_total = sum(section_amounts.values())
        gt_label = ws.cell(row=row + 1, column=1, value="GRAND TOTAL (AED)")
        ws.merge_cells(
            start_row=row + 1, start_column=1,
            end_row=row + 1, end_column=5
        )
        gt_label.fill = _fill(COLOR_HEADER_BG)
        gt_label.font = _font(bold=True, color=COLOR_HEADER_FG, size=12)
        gt_label.alignment = _left()
        gt_label.border = _border()

        gt_value = ws.cell(row=row + 1, column=6, value=grand_total)
        gt_value.fill = _fill(COLOR_HEADER_BG)
        gt_value.font = _font(bold=True, color=COLOR_HEADER_FG, size=12)
        gt_value.number_format = "#,##0.00"
        gt_value.border = _border()

    # ------------------------------------------------------------------
    # Summary Sheet
    # ------------------------------------------------------------------
    def _create_summary_sheet(
        self,
        wb: "Workbook",
        boq: list[dict],
        validation_report: Any,
        info: dict,
    ) -> None:
        ws = wb.create_sheet("Summary", 0)

        row = 1

        # Title
        title_cell = ws.cell(row=row, column=1, value="QTO SYSTEM — BILL OF QUANTITIES SUMMARY")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        title_cell.fill = _fill(COLOR_HEADER_BG)
        title_cell.font = _font(bold=True, color=COLOR_HEADER_FG, size=14)
        title_cell.alignment = _center()
        row += 2

        # Project info
        for label, key in [
            ("Project Name", "name"),
            ("Project Ref", "ref"),
            ("Project Type", "type"),
            ("Plot Area (m²)", "plot_area"),
            ("Date", "date"),
        ]:
            ws.cell(row=row, column=1, value=label).font = _font(bold=True)
            ws.cell(row=row, column=2, value=info.get(key, "N/A"))
            row += 1

        row += 1

        # Overall confidence
        conf = getattr(validation_report, "overall_confidence", 100.0)
        is_draft = getattr(validation_report, "is_draft", True)
        status_label = "DRAFT" if is_draft else "FINAL"
        conf_bg = COLOR_DRAFT_BG if is_draft else COLOR_FINAL_BG

        conf_label = ws.cell(row=row, column=1, value="Overall BOQ Confidence")
        conf_label.font = _font(bold=True, size=12)
        conf_val = ws.cell(row=row, column=2, value=f"{conf:.1f}%  — {status_label}")
        conf_val.fill = _fill(conf_bg)
        conf_val.font = _font(bold=True, size=12)
        row += 1

        summary_text = getattr(validation_report, "summary", "")
        ws.cell(row=row, column=1, value=summary_text)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 2

        # Validation table
        headers = ["Description", "Calculated Qty", "Scaled Average", "Deviation %",
                   "Confidence %", "Flag"]
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.fill = _fill(COLOR_HEADER_BG)
            cell.font = _font(bold=True, color=COLOR_HEADER_FG)
            cell.alignment = _center()
            cell.border = _border()

        ws.column_dimensions["A"].width = 48
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 10
        row += 1

        item_results = getattr(validation_report, "item_results", [])
        for vr in item_results:
            row_vals = [
                vr.item_name,
                vr.calculated_qty,
                vr.scaled_average if vr.scaled_average is not None else "N/A",
                f"{vr.deviation_pct:.1f}%" if vr.deviation_pct is not None else "N/A",
                vr.confidence,
                vr.flag,
            ]
            bg = self._flag_bg(vr.flag)
            fg = self._flag_fg(vr.flag)
            for col_idx, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.border = _border()
                cell.alignment = _center() if col_idx != 1 else _left()
                if col_idx in (5, 6):
                    cell.fill = _fill(bg)
                    cell.font = _font(color=fg)
                else:
                    cell.font = _font()
            row += 1

        # Ratio checks section
        ratio_results = getattr(validation_report, "ratio_results", [])
        if ratio_results:
            row += 1
            ws.cell(row=row, column=1, value="RATIO VALIDATION CHECKS").font = _font(
                bold=True, size=11
            )
            row += 1
            ratio_headers = ["Ratio", "Expected", "Actual", "Deviation %", "Flag", "Note"]
            for col_idx, h in enumerate(ratio_headers, start=1):
                cell = ws.cell(row=row, column=col_idx, value=h)
                cell.fill = _fill(COLOR_HEADER_BG)
                cell.font = _font(bold=True, color=COLOR_HEADER_FG)
                cell.border = _border()
            row += 1
            for rr in ratio_results:
                bg = self._flag_bg(rr.flag)
                fg = self._flag_fg(rr.flag)
                row_vals = [
                    rr.ratio_name, rr.expected_ratio, rr.actual_ratio,
                    f"{rr.deviation_pct:.1f}%", rr.flag, rr.note
                ]
                for col_idx, val in enumerate(row_vals, start=1):
                    cell = ws.cell(row=row, column=col_idx, value=val)
                    cell.border = _border()
                    if col_idx == 5:
                        cell.fill = _fill(bg)
                        cell.font = _font(color=fg)
                    else:
                        cell.font = _font()
                row += 1

    # ------------------------------------------------------------------
    # Analysis Sheet
    # ------------------------------------------------------------------
    def _create_analysis_sheet(
        self,
        wb: "Workbook",
        boq: list[dict],
        info: dict,
    ) -> None:
        """
        Third sheet: cost breakdown, data quality, and top cost drivers.
        Gives management/engineer an instant intelligence snapshot.
        """
        ws = wb.create_sheet("Analysis")

        plot_area: float = float(info.get("plot_area") or 1.0) or 1.0
        grand_total = sum(item.get("amount", 0.0) for item in boq)

        # ---- Title ----
        row = 1
        title = ws.cell(row=row, column=1, value="BOQ COST & DATA QUALITY ANALYSIS")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        title.fill = _fill(COLOR_HEADER_BG)
        title.font = _font(bold=True, color=COLOR_HEADER_FG, size=14)
        title.alignment = _center()
        ws.row_dimensions[row].height = 36
        row += 2

        # ---- Project meta ----
        for label, key in [
            ("Project", "name"), ("Type", "type"),
            ("Plot Area (m²)", "plot_area"), ("Date", "date"),
        ]:
            ws.cell(row=row, column=1, value=label).font = _font(bold=True)
            ws.cell(row=row, column=2, value=info.get(key, "N/A")).font = _font()
            row += 1
        row += 1

        # ---- Data quality summary ----
        total_items = len(boq)
        estimated_items = sum(1 for i in boq if i.get("estimated"))
        calculated_items = total_items - estimated_items
        quality_pct = (calculated_items / total_items * 100.0) if total_items else 0.0

        dq_title = ws.cell(row=row, column=1, value="DATA QUALITY")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        dq_title.fill = _fill(COLOR_SECTION_BG)
        dq_title.font = _font(bold=True, color=COLOR_SECTION_FG, size=11)
        dq_title.alignment = _left()
        row += 1

        dq_rows = [
            ("Total BOQ items",        total_items),
            ("Calculated from drawings", calculated_items),
            ("Estimated from benchmarks", estimated_items),
            ("Data quality",           f"{quality_pct:.1f}%"),
            ("Grand Total Cost (AED)", f"{grand_total:,.2f}"),
            ("Cost per m² (AED/m²)",   f"{grand_total / plot_area:,.2f}" if plot_area else "N/A"),
        ]
        for label, value in dq_rows:
            lbl = ws.cell(row=row, column=1, value=label)
            lbl.font = _font(bold=True)
            lbl.border = _border()
            val = ws.cell(row=row, column=2, value=value)
            val.font = _font()
            val.border = _border()
            if label == "Estimated from benchmarks" and estimated_items > 0:
                val.fill = _fill(COLOR_ESTIMATED_BG)
                val.font = _font(color=COLOR_ESTIMATED_FG)
            elif label == "Data quality":
                val.fill = _fill(COLOR_GREEN_BG if quality_pct >= 80 else
                                 COLOR_YELLOW_BG if quality_pct >= 50 else COLOR_RED_BG)
            row += 1
        row += 1

        # ---- Section cost breakdown ----
        sb_title = ws.cell(row=row, column=1, value="SECTION COST BREAKDOWN")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        sb_title.fill = _fill(COLOR_SECTION_BG)
        sb_title.font = _font(bold=True, color=COLOR_SECTION_FG, size=11)
        sb_title.alignment = _left()
        row += 1

        # Column headers
        sec_headers = ["Section", "Total (AED)", "% of Total", "Cost / m²",
                       "Items", "Calculated", "Estimated"]
        for ci, h in enumerate(sec_headers, start=1):
            c = ws.cell(row=row, column=ci, value=h)
            c.fill = _fill(COLOR_HEADER_BG)
            c.font = _font(bold=True, color=COLOR_HEADER_FG)
            c.alignment = _center()
            c.border = _border()
        row += 1

        # Build section data
        section_data: dict[str, dict] = {}
        for item in boq:
            cat = item.get("category", "General")
            sd = section_data.setdefault(cat, {
                "total": 0.0, "count": 0,
                "calculated": 0, "estimated": 0
            })
            sd["total"] += item.get("amount", 0.0)
            sd["count"] += 1
            if item.get("estimated"):
                sd["estimated"] += 1
            else:
                sd["calculated"] += 1

        for cat, sd in section_data.items():
            pct = (sd["total"] / grand_total * 100.0) if grand_total else 0.0
            cost_pm2 = sd["total"] / plot_area if plot_area else 0.0
            row_vals = [
                cat,
                round(sd["total"], 2),
                f"{pct:.1f}%",
                round(cost_pm2, 2),
                sd["count"],
                sd["calculated"],
                sd["estimated"],
            ]
            for ci, val in enumerate(row_vals, start=1):
                c = ws.cell(row=row, column=ci, value=val)
                c.border = _border()
                c.alignment = _center() if ci != 1 else _left()
                c.font = _font()
                if ci == 2:
                    c.number_format = "#,##0.00"
                if ci == 7 and sd["estimated"] > 0:
                    c.fill = _fill(COLOR_ESTIMATED_BG)
                    c.font = _font(color=COLOR_ESTIMATED_FG)
            row += 1

        # Grand total row
        gt = ws.cell(row=row, column=1, value="GRAND TOTAL")
        gt.fill = _fill(COLOR_HEADER_BG)
        gt.font = _font(bold=True, color=COLOR_HEADER_FG)
        gt.border = _border()
        gt_val = ws.cell(row=row, column=2, value=round(grand_total, 2))
        gt_val.fill = _fill(COLOR_HEADER_BG)
        gt_val.font = _font(bold=True, color=COLOR_HEADER_FG)
        gt_val.number_format = "#,##0.00"
        gt_val.border = _border()
        row += 2

        # ---- Top 10 cost drivers ----
        td_title = ws.cell(row=row, column=1, value="TOP 10 COST DRIVERS")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        td_title.fill = _fill(COLOR_SECTION_BG)
        td_title.font = _font(bold=True, color=COLOR_SECTION_FG, size=11)
        td_title.alignment = _left()
        row += 1

        td_headers = ["Rank", "Description", "Section", "Qty", "Unit",
                      "Amount (AED)", "% of Total"]
        for ci, h in enumerate(td_headers, start=1):
            c = ws.cell(row=row, column=ci, value=h)
            c.fill = _fill(COLOR_HEADER_BG)
            c.font = _font(bold=True, color=COLOR_HEADER_FG)
            c.alignment = _center()
            c.border = _border()
        row += 1

        top_items = sorted(boq, key=lambda x: x.get("amount", 0.0), reverse=True)[:10]
        for rank, item in enumerate(top_items, start=1):
            pct = (item.get("amount", 0.0) / grand_total * 100.0) if grand_total else 0.0
            row_vals = [
                rank,
                item.get("description", ""),
                item.get("category", ""),
                item.get("quantity", 0.0),
                item.get("unit", ""),
                item.get("amount", 0.0),
                f"{pct:.1f}%",
            ]
            flag = item.get("flag", "GREEN")
            bg = self._flag_bg(flag)
            fg = self._flag_fg(flag)
            for ci, val in enumerate(row_vals, start=1):
                c = ws.cell(row=row, column=ci, value=val)
                c.border = _border()
                c.alignment = _center() if ci not in (2, 3) else _left()
                if ci == 6:
                    c.number_format = "#,##0.00"
                    c.font = _font(bold=True)
                elif ci == 1:
                    # Rank coloured by item flag so estimated drivers stand out
                    c.fill = _fill(bg)
                    c.font = _font(bold=True, color=fg)
                else:
                    c.font = _font()
            row += 1

        # ---- Legend ----
        row += 1
        legend_title = ws.cell(row=row, column=1, value="LEGEND")
        legend_title.font = _font(bold=True)
        row += 1
        for flag, label in [
            ("GREEN",     "Calculated — within historical range"),
            ("YELLOW",    "Calculated — moderate deviation from average"),
            ("RED",       "Calculated — high deviation, manual review needed"),
            ("ESTIMATED", "Estimated — qty replaced with scaled historical average"),
        ]:
            bg = self._flag_bg(flag)
            fg = self._flag_fg(flag)
            c = ws.cell(row=row, column=1, value=f"  {flag}")
            c.fill = _fill(bg)
            c.font = _font(bold=True, color=fg)
            c.border = _border()
            lbl = ws.cell(row=row, column=2, value=label)
            lbl.font = _font()
            row += 1

        # ---- Column widths ----
        widths = [28, 14, 12, 13, 8, 16, 10]
        for ci, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(ci)].width = w

    # ------------------------------------------------------------------
    # Header writer
    # ------------------------------------------------------------------
    def _write_project_header(self, ws: Any, info: dict) -> None:
        ws.column_dimensions["A"].width = 10

        title_cell = ws.cell(row=1, column=1,
                             value="BILL OF QUANTITIES — QTO AUTOMATION SYSTEM")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
        title_cell.fill = _fill(COLOR_HEADER_BG)
        title_cell.font = _font(bold=True, color=COLOR_HEADER_FG, size=14)
        title_cell.alignment = _center()
        ws.row_dimensions[1].height = 36

        fields = [
            ("Project:", info.get("name", "N/A")),
            ("Ref:", info.get("ref", "N/A")),
            ("Type:", info.get("type", "N/A")),
            ("Plot Area:", f"{info.get('plot_area', 'N/A')} m²"),
            ("Date:", info.get("date", datetime.today().strftime("%Y-%m-%d"))),
        ]
        col = 1
        for label, value in fields:
            ws.cell(row=2, column=col, value=label).font = _font(bold=True)
            ws.cell(row=2, column=col + 1, value=value).font = _font()
            col += 2

    # ------------------------------------------------------------------
    # Colour helpers
    # ------------------------------------------------------------------
    @staticmethod
    def _flag_bg(flag: str) -> str:
        return {
            "GREEN":     COLOR_GREEN_BG,
            "YELLOW":    COLOR_YELLOW_BG,
            "RED":       COLOR_RED_BG,
            "ESTIMATED": COLOR_ESTIMATED_BG,
        }.get(flag, COLOR_GREEN_BG)

    @staticmethod
    def _flag_fg(flag: str) -> str:
        return {
            "GREEN":     COLOR_GREEN_FG,
            "YELLOW":    COLOR_YELLOW_FG,
            "RED":       COLOR_RED_FG,
            "ESTIMATED": COLOR_ESTIMATED_FG,
        }.get(flag, COLOR_GREEN_FG)

    # ------------------------------------------------------------------
    # Validation map
    # ------------------------------------------------------------------
    @staticmethod
    def _build_validation_map(validation_report: Any) -> dict:
        """Map description → ValidationResult for quick lookup."""
        result: dict = {}
        if validation_report is None:
            return result
        for vr in getattr(validation_report, "item_results", []):
            result[vr.item_name] = vr
        return result
